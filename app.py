# app.py

import streamlit as st
import pandas as pd
import numpy as np
import re
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

# ✅ Page config FIRST
st.set_page_config(page_title="CSI Coder using AI")

# Text cleaning for consistent matching
def clean_text(text):
    text = str(text).lower()
    text = re.sub(r"[\W_]+", " ", text)  # remove punctuation and underscores
    return text.strip()

# Load the hidden CSI reference table (Excel Table named CSI_Integration_Tool)
@st.cache_data
def load_reference():
    file_path = "Master_CSI.xlsx"
    wb = load_workbook(filename=file_path, data_only=True)
    table_name = "CSI_Integration_Tool"
    table_ref = None
    sheet_name = None
    for ws in wb.worksheets:
        if table_name in ws.tables:
            sheet_name = ws.title
            table_ref = ws.tables[table_name].ref
            break
    if not table_ref:
        raise ValueError(f"Table '{table_name}' not found in {file_path}")

    start_col, start_row, end_col, end_row = range_boundaries(table_ref)
    start_col_letter = get_column_letter(start_col)
    end_col_letter = get_column_letter(end_col)
    usecols = f"{start_col_letter}:{end_col_letter}"
    header_row = start_row - 1

    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        header=header_row,
        usecols=usecols,
        engine="openpyxl"
    )
    df = df.iloc[: (end_row - start_row + 1)]
    df = df.dropna(how="all")
    df.columns = [col.strip() for col in df.columns]
    return df

# Load reference data and prepare TF-IDF
reference_df = load_reference()
# Create combined keywords from Division Name & Section 1 Name
combined_kw = (
    reference_df['Division Name'].astype(str) + " " + reference_df['Section 1 Name'].astype(str)
).tolist()
# Clean reference texts
cleaned_ref = [clean_text(t) for t in combined_kw]
# Initialize TF-IDF vectorizer on combined reference texts
vectorizer = TfidfVectorizer().fit(cleaned_ref)
# Transform reference texts into TF-IDF vectors
ref_tfidf = vectorizer.transform(cleaned_ref)

st.title("🧠 CSI Coder using AI")
st.markdown("Batch upload a BOQ file or enter a single item to get CSI codes using TF-IDF matching (fast, no torch).")

batch_tab, single_tab = st.tabs(["Batch Upload","Single Item"])

# Batch upload tab
with batch_tab:
    uploaded_file = st.file_uploader(
        "📎 Upload BOQ Item List (Excel or CSV)",
        type=["xlsx", "csv"]
    )
    if uploaded_file:
        # Auto-detect header row
        if uploaded_file.name.lower().endswith("csv"):
            raw_df = pd.read_csv(uploaded_file, header=None)
        else:
            raw_df = pd.read_excel(uploaded_file, header=None)
        raw_df = raw_df.dropna(how='all')
        headers = raw_df.iloc[0].astype(str).tolist()
        data = raw_df[1:].copy()
        boq_df = data
        boq_df.columns = headers

        selected_col = st.selectbox("🔽 Select the BOQ column to code:", boq_df.columns)
        if st.button("🚀 Match Items to CSI Codes"):
            with st.spinner("Matching using TF-IDF..."):
                raw_texts = boq_df[selected_col].astype(str).tolist()
                cleaned_texts = [clean_text(t) for t in raw_texts]
                # Transform BOQ texts
                text_tfidf = vectorizer.transform(cleaned_texts)

                results = []
                for idx, orig in enumerate(raw_texts):
                    # Compute cosine similarity between single text and all refs
                    scores = cosine_similarity(text_tfidf[idx], ref_tfidf.toarray())[0]
                    top_idx = int(np.argmax(scores))
                    score = float(scores[top_idx])
                    # Determine confidence
                    if score >= 0.5:
                        confidence = 'High'
                    elif score >= 0.2:
                        confidence = 'Medium'
                    else:
                        confidence = 'Low'
                    ref_row = reference_df.iloc[top_idx].to_dict()
                    entry = {"BOQ Description": orig}
                    entry.update(ref_row)
                    entry["Confidence"] = confidence
                    results.append(entry)

                result_df = pd.DataFrame(results)
                cols = ["BOQ Description"] + reference_df.columns.tolist() + ["Confidence"]
                result_df = result_df[cols]

                st.success("✅ Matching complete with confidence levels!")
                st.dataframe(result_df)

                # Download Excel
                output = BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    result_df.to_excel(writer, index=False, sheet_name='Matched Results')
                output.seek(0)
                st.download_button(
                    label="⬇️ Download Matched Results as Excel",
                    data=output,
                    file_name="CSI_Matched_Output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# Single item tab
with single_tab:
    input_text = st.text_input("Or enter a single BOQ item to code:")
    if st.button("🖊️ Code Single Item"):
        if input_text:
            with st.spinner("Matching using TF-IDF..."):
                cleaned_input = clean_text(input_text)
                input_vec = vectorizer.transform([cleaned_input])
                scores = cosine_similarity(input_vec, ref_tfidf.toarray())[0]
                top_idx = int(np.argmax(scores))
                score = float(scores[top_idx])
                if score >= 0.5:
                    confidence = 'High'
                elif score >= 0.2:
                    confidence = 'Medium'
                else:
                    confidence = 'Low'
                ref_row = reference_df.iloc[top_idx].to_dict()
                entry = {"BOQ Description": input_text}
                entry.update(ref_row)
                entry["Confidence"] = confidence
                result_df = pd.DataFrame([entry])
                cols = ["BOQ Description"] + reference_df.columns.tolist() + ["Confidence"]
                result_df = result_df[cols]
                st.success("✅ Single item matched!")
                st.dataframe(result_df)

                # Download single result
                output = BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    result_df.to_excel(writer, index=False, sheet_name='Result')
                output.seek(0)
                st.download_button(
                    label="⬇️ Download Result as Excel",
                    data=output,
                    file_name="CSI_Single_Output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.error("Please enter a BOQ item before coding.")
